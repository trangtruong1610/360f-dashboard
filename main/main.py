from openpyxl import load_workbook

def load_input_from_xlsx(excel):
    wb = load_workbook(excel)
    ws = wb.active

    data = {}
    for a, b in zip(ws['A'], ws['B']):
        data[a.value] = b.value
    return data
