from openpyxl import load_workbook

def load_input_from_xlsx(excel):
    wb = load_workbook(excel)
    ws = wb.active

    data = {}
    for a, b in zip(ws['A'], ws['B']):
        testcase_id, google_keyword = a.value, b.value
        data[testcase_id] = google_keyword
    return data
